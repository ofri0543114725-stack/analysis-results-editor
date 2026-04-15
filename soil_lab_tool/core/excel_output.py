"""
core/excel_output.py
---------------------
Builds a multi-sheet, multi-threshold Hebrew RTL Excel report.

One sheet is created per analysis type found in the records.

Portrait layout (compounds as rows):
  A: compound name | B: CAS | C+: threshold cols | next: יחידות | then: sample cols

Landscape layout (samples as rows — soil gas, many-samples):
  Row headers: compound / CAS / unit | Columns: samples
  Threshold rows appended at bottom.

Colour coding:
  Yellow  — measured value exceeds threshold
  Gray    — below detection limit, but LOD > threshold (uncertain)

Orientation (per sheet):
  Portrait  (compounds as rows)  — when n_compounds >= n_samples
  Landscape (samples as rows)    — when n_samples > n_compounds

Sheet config:
  SOIL_GAS_VOC → "גז קרקע VOC"   µg/m³
  SOIL_VOC     → "קרקע VOC BTEX" mg/kg
  SOIL_TPH     → "קרקע TPH"      mg/kg
  SOIL_METALS  → "קרקע מתכות"    mg/kg DW
  SOIL_PFAS    → "קרקע PFAS"     ng/kg
  GW_VOC       → "מי תהום BTEX"  mg/L
  GW_PFAS      → "מי תהום PFAS"  ng/L
  LOWFLOW      → "pH"             — field parameters, no thresholds
"""

from __future__ import annotations

import io
import math
import os
import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from core.threshold_manager import ThresholdManager, ANALYSIS_THRESHOLDS, THRESHOLD_LABELS


# ── Style constants ───────────────────────────────────────────────────
YELLOW  = PatternFill(start_color="F7C7AC", end_color="F7C7AC", fill_type="solid")
GRAY    = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
ORANGE  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
BLUE_H  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
DARK_H  = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
THIN    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
WRAP_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_L  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
CENTER  = Alignment(horizontal="center", vertical="center")

FHE = {"name": "David",           "size": 9}   # Hebrew / numbers / dates
FEN = {"name": "Times New Roman", "size": 8}   # Pure English text


def _font(val, bold=False) -> Font:
    """
    Choose font based on cell content:
      Pure English (no Hebrew)        → Times New Roman 8
      Hebrew / mixed / numbers / dates → David 9
    """
    s = str(val) if val is not None else ""
    has_hebrew  = any('\u05D0' <= c <= '\u05EA' for c in s)
    has_english = any(c.isalpha() and c.isascii() for c in s)
    # Pure English → Times New Roman 8; everything else (Hebrew, mixed, numbers, dates) → David 9
    base = FEN if (has_english and not has_hebrew) else FHE
    return Font(**base, bold=bold)


def _round_thresh(v) -> float | None:
    """
    Round a threshold value to 2 decimal places for display.

    Examples
    --------
    90.94     → 90.94
    350       → 350.0
    0.45      → 0.45
    0.8       → 0.8
    None      → None
    """
    if v is None or not isinstance(v, (int, float)):
        return v
    return round(v, 2)


def _num_fmt_data(val) -> str:
    """Excel number format for compound data cells.
    Whole numbers use '#,##0' to avoid a trailing decimal point.
    Fractional numbers use '#,##0.###' (up to 3 significant decimal digits).
    """
    if isinstance(val, int):
        return '#,##0'
    if isinstance(val, float) and val % 1 == 0:
        return '#,##0'
    return '#,##0.###'


def _num_fmt_thresh(val) -> str:
    """Excel number format for threshold cells.
    Whole numbers use '#,##0'; fractions use '#,##0.##'.
    """
    if val is None:
        return 'General'
    if isinstance(val, int):
        return '#,##0'
    if isinstance(val, float) and val % 1 == 0:
        return '#,##0'
    return '#,##0.##'


def _fmt_lod(lod: float) -> str:
    """Format a LOD value as a clean string with no trailing zeros or decimal point.

    Examples:
      50.0  → "50"
      0.009 → "0.009"
      0.020 → "0.02"
    Uses up to 3 decimal places of precision.
    """
    if lod == int(lod):
        return str(int(lod))
    return f"{round(lod, 3):.3f}".rstrip("0").rstrip(".")


_DEPTH_PAREN_RE   = re.compile(r'^(.*?)\s*\((\d+\.?\d*)\)\s*$')
# Shimshon-2 "קק-{depth_float}-{borehole_int}[ DUP]" — depth first, borehole last
_DEPTH_FLOAT_BH_RE = re.compile(r'^(.*?)-(\d+\.\d+)-(\d+)(\s+DUP)?\s*$', re.IGNORECASE)
_DEPTH_DOT_RE     = re.compile(r'^(.*?)-(\d+\.\d+)(-DUP)?\s*$', re.IGNORECASE)
_DEPTH_DASHNUM_RE = re.compile(r'^(.*?)\s+-\s+(\d+)-(\d+)(-DUP)?\s*$', re.IGNORECASE)
# Shimshon-1 "קק-{borehole} - {depth_float}[ DUP]" — name SPACE-DASH-SPACE depth with real decimal
_DEPTH_SDS_RE     = re.compile(r'^(.*?)\s+-\s+(\d+\.?\d*)(\s+DUP)?\s*$', re.IGNORECASE)
_DEPTH_SPACE_RE   = re.compile(r'^(.*?)\s+([\d]+(?:\.[\d]+)?)\s*(?:m|מ)?$', re.IGNORECASE)
_WELL_NORM_RE     = re.compile(r'^([\u05D0-\u05EA])(\d)', re.UNICODE)

# Well letter priority for sorting: ק=0, נ=1, others=99
_WELL_LETTER_ORDER: dict[str, int] = {'ק': 0, 'נ': 1}


def _norm_borehole(s: str) -> str:
    """ק12 → ק-12, נ1 → נ-1 (add '-' after single Hebrew letter prefix).
    Also strips trailing dashes that appear in some lab report formats (e.g. 'קק-3-')."""
    normalized = _WELL_NORM_RE.sub(r'\1-\2', s.strip())
    return normalized.rstrip('-').strip()


def _borehole_sort_key(bh: str) -> tuple:
    """Sort: ק-* first, then נ-*, then others. Within each group: numeric order."""
    bh_n = _norm_borehole(bh)
    first = bh_n[0] if bh_n else ''
    priority = _WELL_LETTER_ORDER.get(first, 99)
    m = re.search(r'(\d+)', bh_n)
    num = int(m.group(1)) if m else 0
    return (priority, num, bh_n)


def _dup_rich_text(bh: str):
    """
    Return CellRichText for borehole names containing 'DUP':
      Hebrew/numbers/punctuation → David 9
      'DUP' → Times New Roman 8
    Returns plain str when no DUP present.
    """
    m = re.search(r'(DUP)', bh, re.IGNORECASE)
    if not m:
        return bh
    he_if = InlineFont(rFont="David", sz=9)
    en_if = InlineFont(rFont="Times New Roman", sz=8)
    parts = []
    before = bh[:m.start()]
    after  = bh[m.end():]
    if before:
        parts.append(TextBlock(he_if, before))
    parts.append(TextBlock(en_if, m.group(1)))
    if after:
        parts.append(TextBlock(he_if, after))
    return CellRichText(*parts)


def _mixed_rich_text(s: str, bold: bool = False):
    """
    For mixed Hebrew+English strings (e.g. 'VSL קרקע'):
      English segments → Times New Roman 8
      Hebrew/other segments → David 9
    Returns CellRichText if mixed, else plain str.
    """
    has_heb = any('\u05D0' <= c <= '\u05EA' for c in s)
    has_eng = any(c.isalpha() and c.isascii() for c in s)
    if not (has_heb and has_eng):
        return s

    he_if = InlineFont(rFont="David", sz=9, b=bold)
    en_if = InlineFont(rFont="Times New Roman", sz=8, b=bold)

    segments: list[tuple[bool, str]] = []  # (is_hebrew, text)
    cur_text = ""
    cur_heb: bool | None = None

    for ch in s:
        is_heb = '\u05D0' <= ch <= '\u05EA'
        is_eng = ch.isalpha() and ch.isascii()
        ch_type = True if is_heb else (False if is_eng else None)  # None = neutral

        if ch_type is None:
            cur_text += ch
        elif ch_type != cur_heb and cur_heb is not None:
            segments.append((cur_heb, cur_text))
            cur_text = ch
            cur_heb = ch_type
        else:
            cur_text += ch
            cur_heb = ch_type

    if cur_text:
        segments.append((cur_heb if cur_heb is not None else False, cur_text))

    if len(segments) <= 1:
        return s
    return CellRichText(*[
        TextBlock(he_if if is_heb else en_if, txt) for is_heb, txt in segments
    ])


def _split_sample_depth(sid: str) -> tuple[str, str]:
    """
    Split sample ID into (borehole_name, depth_str). Handles formats:
      'ק16 (3.0)'        → ('ק-16', '3.0')
      'ק17  DUP(1.2)'    → ('ק-17 DUP', '1.2')
      'ק-16-1.2'         → ('ק-16', '1.2')
      'ק-16-1.2-DUP'     → ('ק-16 DUP', '1.2')
      'ק-16 - 1-2'       → ('ק-16', '1.2')
      'ק-16 - 1-2-DUP'   → ('ק-16 DUP', '1.2')
      'ב-1 3.0'          → ('ב-1', '3.0')
      'קק-1 - 1.5'       → ('קק-1', '1.5')   ← shimshon-1
      'קק-1.5-10'        → ('קק-10', '1.5')  ← shimshon-2
      'קק-10.0-16'       → ('קק-16', '10.0') ← shimshon-2
      'קק-3.0-14 DUP'    → ('קק-14 DUP', '3.0') ← shimshon-2 DUP
    Borehole name is always normalized (ק12 → ק-12).
    """
    s = sid.strip()
    # "name (depth)" or "name DUP(depth)"
    m = _DEPTH_PAREN_RE.match(s)
    if m:
        return _norm_borehole(m.group(1).strip()), m.group(2)

    # Shimshon-2: "prefix-depth_float-borehole_int[ DUP]"
    # e.g. "קק-1.5-10" → borehole="קק-10", depth="1.5"
    m = _DEPTH_FLOAT_BH_RE.match(s)
    if m:
        prefix   = _norm_borehole(m.group(1).strip())
        depth    = m.group(2)
        bh_num   = m.group(3)
        dup_sfx  = " DUP" if m.group(4) else ""
        borehole = f"{prefix}-{bh_num}{dup_sfx}"
        return _norm_borehole(borehole), depth

    # "name-depth[-DUP]" with real decimal point (and no trailing integer segment)
    m = _DEPTH_DOT_RE.match(s)
    if m:
        name = _norm_borehole(m.group(1).strip())
        return (name + ' DUP' if m.group(3) else name), m.group(2)

    # "name - d-d[-DUP]" dash-as-decimal (e.g. "1-2" = 1.2 m)
    m = _DEPTH_DASHNUM_RE.match(s)
    if m:
        name  = _norm_borehole(m.group(1).strip())
        depth = f"{m.group(2)}.{m.group(3)}"
        return (name + ' DUP' if m.group(4) else name), depth

    # Shimshon-1: "name - depth[ DUP]" with space-dash-space and real decimal
    # e.g. "קק-1 - 1.5" → ('קק-1', '1.5')
    # e.g. "קק-3 - 3.0 DUP" → ('קק-3 DUP', '3.0')
    m = _DEPTH_SDS_RE.match(s)
    if m:
        name = _norm_borehole(m.group(1).strip())
        if m.group(3):
            name += " DUP"
        return name, m.group(2)

    # "name depth[m]" space-separated
    m = _DEPTH_SPACE_RE.match(s)
    if m:
        return _norm_borehole(m.group(1).strip()), m.group(2).strip()
    return _norm_borehole(s), ""


# ── Threshold source footnote labels ─────────────────────────────────
_THRESHOLD_SOURCES: dict[str, str] = {
    "GAS_INDOOR_RES":        "Tier 1 RBTL Residential, Rev.7, 12/24",
    "GAS_OUTDOOR_RES":       "Tier 1 RBTL Residential, Rev.7, 12/24",
    "GAS_INDOOR_IND":        "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "GAS_OUTDOOR_IND":       "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_RES_SOIL_VH":     "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_HM_0_6": "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_HM_6":   "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_RES_SOIL_LOW":    "Tier 1 RBTL Residential, Rev.7, 12/24",
    "TIER1_IND_SOIL_VH":     "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_HM_0_6": "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_HM_6":   "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "TIER1_IND_SOIL_LOW":    "Tier 1 RBTL Industrial/Commercial, Rev.7, 12/24",
    "VSL_SOIL":              "Soil VSL, Rev. 7, 12/24",
    "GW":                    "Groundwater Standard, Rev.7, 12/24",
    "PFAS_VSL":              "PFAS VSL, Rev.7, 12/24",
    "PFAS_TIER1_RES":        "PFAS Tier 1 Residential, Rev.7, 12/24",
    "PFAS_TIER1_IND":        "PFAS Tier 1 Industrial/Commercial, Rev.7, 12/24",
}

# ── Sheet configuration ───────────────────────────────────────────────
SHEET_CONFIG: dict[str, dict] = {
    # include_lod_loq  → add LOD + LOQ columns between CAS and threshold(s)
    # filter_nd_safe   → exclude compounds that are ND everywhere AND LOD ≤ threshold
    # units_in_header  → embed unit in column headers; no separate יחידות column
    "SOIL_GAS_VOC": {
        "name": "גז קרקע VOC", "unit": "µg/m³",
        "include_lod_loq": True,
        "filter_nd_safe":  False,
        "units_in_header": True,
    },
    "SOIL_VOC":       {"name": "קרקע VOCs",          "unit": "mg/kg"},
    "SOIL_MBTEX":     {"name": "קרקע MBTEX",         "unit": "mg/kg"},
    "SOIL_TPH":       {"name": "קרקע TPH",           "unit": "mg/kg"},
    "SOIL_TPH_VOC":   {"name": "קרקע TPH+BTEX",      "unit": "mg/kg"},
    "SOIL_TPH_MBTEX": {"name": "קרקע TPH+MBTEX",     "unit": "mg/kg"},
    "SOIL_METALS":    {"name": "קרקע מתכות",         "unit": "mg/kg DW"},
    "SOIL_PFAS":   {"name": "קרקע PFAS",       "unit": "ng/kg"},
    "GW_VOC":      {"name": "מי תהום BTEX",    "unit": "mg/L"},
    "GW_PFAS":     {"name": "מי תהום PFAS",    "unit": "ng/L"},
    "LOWFLOW":     {"name": "pH",               "unit": ""},
}


def _ordered_unique(seq) -> list:
    seen = set()
    out = []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


class LabReportExcel:
    """
    Build a multi-sheet Hebrew RTL Excel lab report.

    Parameters
    ----------
    records : list[dict]
        Flat list of measurement records, each with keys:
        compound, cas, sample_id, value, flag, unit, lod, analysis_type
    threshold_manager : ThresholdManager
    output_path : str
    project_name : str
    client : str
    report_date : str   (DD.MM.YYYY)
    selected_thresholds : list[str] | None
        Override which threshold keys to show.  None = use defaults per analysis.
    """

    def __init__(
        self,
        records: list[dict],
        threshold_manager: ThresholdManager,
        output_path: str = "lab_report.xlsx",
        project_name: str = "",
        client: str = "",
        report_date: str = "",
        selected_thresholds: list[str] | None = None,
        combine_tph_voc: bool = False,
        combine_tph_mbtex: bool = False,
    ):
        self.records           = records
        self.tm                = threshold_manager
        self.out_path          = output_path
        self.project           = project_name
        self.client            = client
        self.rep_date          = report_date or date.today().strftime("%d.%m.%Y")
        self.sel_thresh        = selected_thresholds  # None → auto per analysis_type
        self.combine_tph_voc   = combine_tph_voc
        self.combine_tph_mbtex = combine_tph_mbtex

    # ------------------------------------------------------------------
    def build(self) -> str:
        # Group records by analysis_type
        groups: dict[str, list[dict]] = defaultdict(list)
        for r in self.records:
            groups[r.get("analysis_type", "UNKNOWN")].append(r)

        # Optionally merge SOIL_TPH + SOIL_VOC into one combined sheet
        if self.combine_tph_voc and "SOIL_TPH" in groups and "SOIL_VOC" in groups:
            groups["SOIL_TPH_VOC"] = list(groups.pop("SOIL_TPH")) + list(groups.pop("SOIL_VOC"))

        # Optionally merge SOIL_TPH + SOIL_MBTEX into one combined sheet
        if self.combine_tph_mbtex and "SOIL_TPH" in groups and "SOIL_MBTEX" in groups:
            groups["SOIL_TPH_MBTEX"] = list(groups.pop("SOIL_TPH")) + list(groups.pop("SOIL_MBTEX"))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default sheet

        for atype, recs in groups.items():
            cfg   = SHEET_CONFIG.get(atype, {"name": atype, "unit": ""})
            sheet = wb.create_sheet(title=cfg["name"][:31])
            sheet.sheet_view.rightToLeft = True

            thresh_keys = self._thresh_keys(atype)

            if atype == "LOWFLOW":
                self._write_lowflow_sheet(sheet, recs, cfg)
            else:
                self._write_data_sheet(sheet, recs, cfg, thresh_keys)

        # Only create directories when out_path is a real filesystem path (not BytesIO)
        if isinstance(self.out_path, (str, os.PathLike)):
            os.makedirs(os.path.dirname(self.out_path) or ".", exist_ok=True)
        wb.save(self.out_path)
        return self.out_path

    # ------------------------------------------------------------------
    # Sheet writers
    # ------------------------------------------------------------------
    def _write_data_sheet(self, ws, records, cfg, thresh_keys):
        samples   = _ordered_unique(r["sample_id"] for r in records)
        compounds = _ordered_unique(r["compound"]  for r in records)

        # Pivot: compound → sample_id → (value, flag, lod)
        pivot:    dict[str, dict] = {}
        cas_map:  dict[str, str]  = {}
        lod_map:  dict[str, float | None] = {}
        loq_map:  dict[str, float | None] = {}
        unit_map: dict[str, str]  = {}
        for r in records:
            cmp = r["compound"]
            sid = r["sample_id"]
            if cmp not in pivot:
                pivot[cmp]    = {}
                cas_map[cmp]  = r.get("cas", "")
                lod_map[cmp]  = r.get("lod")
                loq_map[cmp]  = r.get("loq")
                unit_map[cmp] = r.get("unit", cfg.get("unit", ""))
            pivot[cmp][sid] = (r.get("value"), r.get("flag", ""), r.get("lod"))

        # Get thresholds per compound
        thresh_vals: dict[str, dict[str, float | None]] = {}
        for cmp, cas in cas_map.items():
            thresh_vals[cmp] = {
                k: self.tm.get_threshold_with_name(cas, k, compound_name=cmp)
                for k in thresh_keys
            }

        # Optional: remove compounds that are ND everywhere AND LOD ≤ strictest threshold
        # (safe to exclude — cannot possibly exceed threshold)
        if cfg.get("filter_nd_safe"):
            def _should_keep(cmp: str) -> bool:
                t_limit = self._strictest(thresh_vals.get(cmp, {}))
                for sid in samples:
                    v, flag, lod = pivot.get(cmp, {}).get(sid, (None, "ND", None))
                    # At least one detected value → keep
                    if flag not in ("ND", "<LOQ") and v is not None:
                        return True
                    # ND but LOD exceeds threshold → grey → keep
                    if lod is not None and t_limit is not None and lod > t_limit:
                        return True
                return False
            compounds = [c for c in compounds if _should_keep(c)]

        # Per-sample metadata (soil gas: canister, sampling date, PID reading)
        sample_meta: dict[str, dict] = {}
        for r in records:
            sid = r["sample_id"]
            if sid not in sample_meta:
                sample_meta[sid] = {
                    "canister": r.get("canister_num", ""),
                    "date":     r.get("sampling_date", ""),
                    "pid":      r.get("pid_reading", ""),
                }

        # Decide orientation: portrait when n_compounds >= n_samples
        portrait = len(compounds) >= len(samples)

        header_info = {
            "project": self.project,
            "date":    self.rep_date,
            "client":  self.client,
            "unit":    cfg["unit"],
        }

        if portrait:
            self._write_portrait(ws, compounds, samples, pivot, cas_map,
                                 lod_map, loq_map,
                                 thresh_keys, thresh_vals, header_info, cfg,
                                 sample_meta=sample_meta, unit_map=unit_map)
        else:
            self._write_landscape(ws, compounds, samples, pivot, cas_map,
                                  lod_map, loq_map,
                                  thresh_keys, thresh_vals, header_info, cfg,
                                  sample_meta=sample_meta, unit_map=unit_map)

    def _write_lowflow_sheet(self, ws, records, cfg):
        """LOWFLOW/pH: field parameters as rows, samples as columns, no thresholds.
        Extracts borehole name and depth from sample IDs (rows 2-3 metadata).
        """
        samples = _ordered_unique(r["sample_id"] for r in records)
        params  = _ordered_unique(r["compound"]  for r in records)

        pivot: dict[str, dict] = {}
        unit_map: dict[str, str] = {}
        for r in records:
            p = r["compound"]
            s = r["sample_id"]
            if p not in pivot:
                pivot[p]    = {}
                unit_map[p] = r.get("unit", "")
            v = r.get("value")
            pivot[p][s] = (round(v, 3) if isinstance(v, (int, float)) else v,
                           r.get("flag", ""), None)

        # Split sample IDs into borehole + depth
        split_map  = {sid: _split_sample_depth(sid) for sid in samples}
        boreholes  = [split_map[sid][0] for sid in samples]
        depths     = [split_map[sid][1] for sid in samples]

        N_FIXED = 2   # פרמטר | יחידות
        total_cols = N_FIXED + len(samples)

        # Row 1: merged project header
        self._write_header_row(ws, 1, total_cols)

        # Rows 2-3: metadata (שם קידוח, עומק [מ'])
        meta_rows = [
            ("שם קידוח",  boreholes),
            ("עומק [מ']", depths),
        ]
        for ri, (label, vals) in enumerate(meta_rows, 2):
            ws.merge_cells(start_row=ri, start_column=1,
                           end_row=ri,   end_column=N_FIXED)
            c = ws.cell(row=ri, column=1, value=label)
            c.font      = Font(**FHE, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN
            ws.cell(row=ri, column=2).border = THIN
            for ci, v in enumerate(vals, N_FIXED + 1):
                cell = ws.cell(row=ri, column=ci, value=v if v else "")
                cell.border    = THIN
                cell.alignment = CENTER
                cell.font      = _font(v)

        # Row 4: column headers (no fill — rows 1-4 are fill-free)
        hdr_row = 4
        headers = ["פרמטר", "יחידות"] + samples
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=hdr_row, column=ci, value=h)
            c.font      = Font(**FHE, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN

        # Data rows
        row_num = hdr_row + 1
        for param in params:
            row_data = [param, unit_map.get(param, "")]
            for sid in samples:
                v, flag, _ = pivot.get(param, {}).get(sid, (None, "ND", None))
                row_data.append(v if v is not None else "N.D.")
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font      = _font(val)
                c.alignment = CENTER
                c.border    = THIN
            row_num += 1

        # Note
        note = ws.cell(row=row_num + 1, column=1,
                       value="* ממצאי שדה בלבד, ללא השוואה לערכי סף")
        note.font = Font(**FHE, italic=True, color="808080")

        self._auto_width(ws, total_cols, hdr_row=4)

    # ------------------------------------------------------------------
    # Portrait layout: compounds as rows
    #
    # Standard mode (soil / GW):
    #   A: compound | B: CAS | C…: threshold col(s) | next: יחידות | then: samples
    #
    # LOD/LOQ mode (soil-gas, cfg["include_lod_loq"]=True):
    #   A: compound | B: CAS | C: LOD [unit] | D: LOQ [unit]
    #   | E…: threshold col(s) | then: samples  (no separate יחידות col)
    # ------------------------------------------------------------------
    def _write_portrait(self, ws, compounds, samples, pivot, cas_map,
                        lod_map, loq_map,
                        thresh_keys, thresh_vals, hinfo, cfg=None, sample_meta=None,
                        unit_map=None):
        cfg         = cfg or {}
        sample_meta = sample_meta or {}
        unit            = hinfo["unit"]
        include_lod_loq = cfg.get("include_lod_loq", False)
        units_in_header = cfg.get("units_in_header", False)

        N_COMPOUND = 2                         # A: compound, B: CAS Number
        N_LOD_LOQ  = 2 if include_lod_loq else 0
        N_THRESH   = len(thresh_keys)
        N_UNIT     = 0                             # unit shown in "Final conc." header
        N_FIXED    = N_COMPOUND + N_LOD_LOQ + N_THRESH + N_UNIT
        total_cols = N_FIXED + len(samples)

        thresh_labels = [THRESHOLD_LABELS.get(k, k) for k in thresh_keys]

        if include_lod_loq:
            # ── Rows 1-4: sample metadata rows (soil gas) ─────────────
            meta_rows = [
                ("שם קידוח",                        [s for s in samples]),
                ("תאריך ביצוע הדיגום",              [sample_meta.get(s, {}).get("date",     "") for s in samples]),
                ("מספר קניסטר",                     [sample_meta.get(s, {}).get("canister", "") for s in samples]),
                ('קריאת PID בסיום השאיבה [חל"מ]',  [sample_meta.get(s, {}).get("pid",      "") for s in samples]),
            ]
            for ri, (label, vals) in enumerate(meta_rows, 1):
                # Merge label across all fixed columns (A → last fixed col)
                ws.merge_cells(start_row=ri, start_column=1,
                               end_row=ri,   end_column=N_FIXED)
                if ri == 4:
                    # Rich text: Hebrew parts → David 9 bold; "PID" → Times New Roman 8 bold
                    he_if = InlineFont(rFont="David", sz=9, b=True)
                    en_if = InlineFont(rFont="Times New Roman", sz=8, b=True)
                    parts = label.split("PID")
                    c = ws.cell(row=ri, column=1)
                    c.value = CellRichText(
                        TextBlock(he_if, parts[0]),
                        TextBlock(en_if, "PID"),
                        TextBlock(he_if, parts[1]),
                    )
                    c.font = Font(**FHE, bold=True)
                else:
                    c = ws.cell(row=ri, column=1, value=label)
                    c.font = _font(label, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
                # No fill on label or fixed columns
                for ci in range(2, N_FIXED + 1):
                    ws.cell(row=ri, column=ci).border = THIN
                # Sample value cells
                is_date_row = (ri == 2)   # row 2 = תאריך ביצוע הדיגום
                for ci, v in enumerate(vals, N_FIXED + 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.border    = THIN
                    cell.alignment = CENTER
                    if is_date_row:
                        # Leave empty for manual entry; apply date format
                        cell.number_format = "DD/MM/YYYY"
                    else:
                        cell.value = v
                        cell.font  = _font(v)
            # ── Row 5: column headers (no fill; sample cols merged) ────
            lod_hdr = f"LOD [{unit}]"
            loq_hdr = f"LOQ [{unit}]"
            headers = (["תרכובת", "CAS Number", lod_hdr, loq_hdr]
                       + thresh_labels
                       + [""] * len(samples))
            hdr_row = 5
        else:
            # ── Row 1: merged project info header ─────────────────────
            self._write_header_row(ws, 1, total_cols, hinfo)
            # ── Rows 2-4: sample metadata ──────────────────────────────
            # Sort samples: ק first, נ second, others; within group by number then depth
            split_p = {sid: _split_sample_depth(sid) for sid in samples}
            samples = sorted(samples,
                             key=lambda sid: (*_borehole_sort_key(split_p[sid][0]),
                                              float(split_p[sid][1]) if split_p[sid][1] else 0.0))
            boreholes = [_dup_rich_text(split_p[sid][0]) for sid in samples]
            depths    = [split_p[sid][1]                 for sid in samples]
            meta_rows = [
                ("שם קידוח",        boreholes),
                ("עומק [מ']",       depths),
                ("קריאת PID [ppm]", [""] * len(samples)),
            ]
            for ri, (label, vals) in enumerate(meta_rows, 2):
                ws.merge_cells(start_row=ri, start_column=1,
                               end_row=ri,   end_column=N_FIXED)
                c = ws.cell(row=ri, column=1, value=label)
                c.font      = _font(label, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
                for ci in range(2, N_FIXED + 1):
                    ws.cell(row=ri, column=ci).border = THIN
                for ci, v in enumerate(vals, N_FIXED + 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.border    = THIN
                    cell.alignment = CENTER
                    if v:
                        cell.value = v
                        if not isinstance(v, CellRichText):
                            cell.font = _font(v)
            # ── Row 5: column headers ──────────────────────────────────
            headers = (["תרכובת", "CAS Number"]
                       + thresh_labels
                       + samples)
            hdr_row = 5

        # Write fixed column headers (no fill on any)
        for ci, h in enumerate(headers[:N_FIXED], 1):
            rv = _mixed_rich_text(h, bold=True) if isinstance(h, str) else h
            c = ws.cell(row=hdr_row, column=ci, value=rv)
            c.font      = _font(h, bold=True)
            c.alignment = WRAP_C
            c.border    = THIN

        # Merge sample columns in header row → "final conc. [unit]"
        if len(samples) > 0:
            sample_start = N_FIXED + 1
            sample_end   = N_FIXED + len(samples)
            if sample_end > sample_start:
                ws.merge_cells(start_row=hdr_row, start_column=sample_start,
                               end_row=hdr_row,   end_column=sample_end)
            conc_hdr = ws.cell(row=hdr_row, column=sample_start,
                               value=f"Final conc. [{unit}]")
            conc_hdr.font      = _font(f"Final conc. [{unit}]", bold=True)
            conc_hdr.alignment = CENTER
            conc_hdr.border    = THIN
            for ci in range(sample_start + 1, sample_end + 1):
                ws.cell(row=hdr_row, column=ci).border = THIN

        # ── Data rows ─────────────────────────────────────────────────
        data_row = hdr_row + 1
        has_gray = False   # tracks whether any gray-filled cell was written

        for cmp in compounds:
            cas    = cas_map.get(cmp, "")
            t_vals = thresh_vals.get(cmp, {})

            # LOD / LOQ (rounded to 3 dp for display — preserves values like 0.009)
            lod_val = lod_map.get(cmp)
            loq_val = loq_map.get(cmp)
            lod_disp = round(lod_val, 3) if isinstance(lod_val, float) else lod_val
            loq_disp = round(loq_val, 3) if isinstance(loq_val, float) else loq_val

            # Threshold values (one per selected threshold key)
            thresh_row = [
                _round_thresh(t_vals.get(k)) if _round_thresh(t_vals.get(k)) is not None
                else "לא קיים"
                for k in thresh_keys
            ]

            # Sample values — build display strings + keep raw for colouring
            sample_vals: list = []
            for sid in samples:
                v, flag, lod = pivot.get(cmp, {}).get(sid, (None, "ND", None))
                if flag == "ND" or (v is None and flag not in ("<LOD", "<LOQ")):
                    # Not detected → show LOD number when available, else "ND"
                    display = round(lod, 3) if lod is not None else "ND"
                elif flag == "<LOD":
                    # <DL / <MDL / <LOD in input → <actual_lod_number (no trailing .0)
                    display = f"<{_fmt_lod(lod)}" if lod is not None else "ND"
                elif flag == "<LOQ":
                    # <LOQ → plain LOQ number (no < prefix)
                    loq_ref = loq_val or v
                    display = round(loq_ref, 3) if isinstance(loq_ref, float) else loq_ref
                elif flag == "<":
                    # Explicit <numeric in input → keep < prefix
                    display = f"<{round(v, 2)}" if isinstance(v, float) else f"<{v}"
                else:
                    display = round(v, 2) if isinstance(v, float) else v
                sample_vals.append((display, v, flag, lod))

            if include_lod_loq:
                row_data = ([cmp, cas, lod_disp, loq_disp]
                            + thresh_row
                            + [sv[0] for sv in sample_vals])
            else:
                row_data = ([cmp, cas]
                            + thresh_row
                            + [sv[0] for sv in sample_vals])

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                c.font      = _font(val)
                c.alignment = WRAP_C if ci == 1 else CENTER
                c.border    = THIN

                # ── LOD / LOQ columns ──────────────────────────────────
                if include_lod_loq and N_COMPOUND < ci <= N_COMPOUND + N_LOD_LOQ:
                    pass   # no special fill, just left as-is

                # ── Threshold columns: no fill ─────────────────────────
                elif (N_COMPOUND + N_LOD_LOQ) < ci <= (N_COMPOUND + N_LOD_LOQ + N_THRESH):
                    if val == "לא קיים":
                        c.font = Font(**FHE, color="808080", italic=True)
                    else:
                        c.font          = Font(**FHE)   # David 9, no bold
                        c.number_format = _num_fmt_thresh(val)
                        c.alignment     = CENTER

                # ── Sample columns: colour coding ──────────────────────
                elif ci > N_FIXED:
                    si = ci - N_FIXED - 1
                    display, num_v, flag, lod = sample_vals[si]
                    thresh_limit = self._strictest(t_vals)
                    if thresh_limit is not None:
                        # YELLOW: actual detected value exceeds threshold
                        if (flag not in ("ND", "<LOQ", "<")
                                and isinstance(num_v, (int, float))
                                and num_v > thresh_limit):
                            c.fill = ORANGE
                            c.font = Font(**FHE, bold=True)
                        # GREY + BOLD: threshold < LOD → false positive risk
                        elif flag in ("ND", "<LOD", "<LOQ", "<"):
                            lod_num = (lod if lod is not None
                                       else (num_v if isinstance(num_v, (int, float)) else None))
                            if lod_num is not None and lod_num > thresh_limit:
                                c.fill = GRAY
                                c.font = _font(display, bold=True)
                                has_gray = True

            data_row += 1

        # ── Legend ────────────────────────────────────────────────────
        self._write_legend(ws, data_row + 1, include_gray=has_gray)
        # ── Threshold source footnotes (only for keys with ≥1 defined value) ──
        active_keys = [k for k in thresh_keys
                       if any(thresh_vals.get(c, {}).get(k) is not None for c in compounds)]
        note_row = data_row + 2
        for note in self._threshold_source_notes(active_keys):
            ws.cell(row=note_row, column=1, value=f"* {note}").font = Font(
                **FEN, italic=True, color="808080")
            note_row += 1
        if include_lod_loq:
            ws.cell(row=note_row, column=1,
                    value="* ספי חש מוגדרים לפי תקנות איכות אויר").font = Font(
                        **FHE, italic=True, color="808080")
        self._auto_width(ws, N_FIXED + len(samples), hdr_row=5)

    # ------------------------------------------------------------------
    # Landscape layout: samples as rows (when n_samples > n_compounds)
    # ------------------------------------------------------------------
    def _write_landscape(self, ws, compounds, samples, pivot, cas_map,
                         lod_map, loq_map,
                         thresh_keys, thresh_vals, hinfo, cfg=None, sample_meta=None,
                         unit_map=None):
        cfg      = cfg or {}
        unit_map = unit_map or {}

        # ── Depth detection & sample sorting ────────────────────────────
        # Split each sample_id into (borehole, depth_str) for display.
        # If ANY sample has a depth suffix (e.g. "ב-1 3.0") we add a depth column.
        split_map = {sid: _split_sample_depth(sid) for sid in samples}
        has_depth = any(depth for _, depth in split_map.values())

        if has_depth:
            # Sort samples: ק first, נ second, others last; within group by number then depth
            def _sort_key(sid):
                bh, dep = split_map[sid]
                return (*_borehole_sort_key(bh), float(dep) if dep else 0.0)
            samples = sorted(samples, key=_sort_key)

        # Column count: borehole + depth (when present) + PID + compounds
        depth_offset  = 1 if has_depth else 0
        PID_COL       = 1 + depth_offset + 1          # 1-based index of PID column
        total_cols    = 1 + depth_offset + 1 + len(compounds)   # +1 for PID
        cmp_col_start = 2 + depth_offset + 1          # 1-based col of first compound

        # ── Row 1: merged project header ────────────────────────────────
        self._write_header_row(ws, 1, total_cols, hinfo)

        # ── Rows 2-4: compound names / CAS / unit ───────────────────────
        if has_depth:
            row2_data = ["שם קידוח", "עומק [מ']", "PID [ppm]"] + compounds
            row3_data = ["CAS Number", "", ""]  + [cas_map.get(c, "") for c in compounds]
            row4_data = ["יחידות", "", ""]      + [unit_map.get(c, hinfo["unit"]) for c in compounds]
        else:
            row2_data = ["שם קידוח", "PID [ppm]"] + compounds
            row3_data = ["CAS Number", ""]  + [cas_map.get(c, "") for c in compounds]
            row4_data = ["יחידות", ""]      + [unit_map.get(c, hinfo["unit"]) for c in compounds]

        for ri, row_vals in enumerate([row2_data, row3_data, row4_data], 2):
            for ci, v in enumerate(row_vals, 1):
                rv = _mixed_rich_text(v, bold=True) if isinstance(v, str) else v
                c = ws.cell(row=ri, column=ci, value=rv)
                c.font      = _font(v, bold=True)
                c.alignment = WRAP_C
                c.border    = THIN
                # No fill on header rows 2-4 (rows 1-4 are fill-free)

        # ── Rows 5+: threshold rows (BEFORE sample data) ────────────────
        UNDEF_FONT  = Font(**FHE, color="808080", italic=True)

        data_row = 5
        for tk in thresh_keys:
            label    = THRESHOLD_LABELS.get(tk, tk)
            # Use plain string with readingOrder=2 so Excel treats the paragraph as RTL
            # even when the label starts with an English acronym like "TIER1".
            lbl_cell = ws.cell(row=data_row, column=1, value=label)
            lbl_cell.font      = _font(label, bold=False)
            lbl_cell.border    = THIN
            lbl_cell.alignment = Alignment(horizontal="right", vertical="center",
                                           wrap_text=True, readingOrder=2)
            # Fill fixed cols (depth + PID) in threshold rows
            for fc in range(2, cmp_col_start):
                ws.cell(row=data_row, column=fc).border = THIN
            for ci, cmp in enumerate(compounds, cmp_col_start):
                cas  = cas_map.get(cmp, "")
                tval = _round_thresh(self.tm.get_threshold(cas, tk))
                c = ws.cell(row=data_row, column=ci)
                c.border = THIN
                if tval is None:
                    c.value     = "לא קיים"
                    c.font      = UNDEF_FONT
                    c.alignment = CENTER
                else:
                    c.value         = tval
                    c.font          = Font(**FHE)   # David 9, not bold
                    c.number_format = _num_fmt_thresh(tval)
                    c.alignment     = CENTER
            data_row += 1

        # ── Sample data rows ─────────────────────────────────────────────
        first_sample_row = data_row   # remember for borehole-merge pass
        has_gray = False              # tracks whether any gray-filled cell was written

        for sid in samples:
            borehole, depth_str = split_map[sid]
            row_meta: list[tuple] = []
            col_vals: list = []
            bh_cell_val = _dup_rich_text(borehole)  # rich text if DUP, else plain

            if has_depth:
                col_vals = [bh_cell_val, depth_str if depth_str else "", ""]  # empty PID
            else:
                col_vals = [bh_cell_val, ""]  # borehole + empty PID

            for cmp in compounds:
                v, flag, lod = pivot.get(cmp, {}).get(sid, (None, "ND", None))
                loq_val = loq_map.get(cmp)
                if flag == "ND" or (v is None and flag not in ("<LOD", "<LOQ")):
                    # Not detected → show LOD number when available, else "ND"
                    display = round(lod, 3) if lod is not None else "ND"
                elif flag == "<LOD":
                    # <DL / <MDL / <LOD in input → <actual_lod_number (no trailing .0)
                    display = f"<{_fmt_lod(lod)}" if lod is not None else "ND"
                elif flag == "<LOQ":
                    # <LOQ → plain LOQ number (no < prefix)
                    loq_ref = loq_val or v
                    display = round(loq_ref, 3) if isinstance(loq_ref, float) else loq_ref
                elif flag == "<":
                    # Explicit <numeric in input → keep < prefix
                    loq_ref = loq_val or v
                    display = (f"<{round(loq_ref, 2)}"
                               if isinstance(loq_ref, float) else f"<{loq_ref}")
                else:
                    display = round(v, 2) if isinstance(v, float) else v
                col_vals.append(display)
                row_meta.append((v, flag, lod))

            for ci, val in enumerate(col_vals, 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                # CellRichText carries its own fonts; plain values use _font()
                if not isinstance(val, CellRichText):
                    c.font = _font(val)
                # Thousands-separator for numeric compound values
                if ci >= cmp_col_start and isinstance(val, (int, float)):
                    c.number_format = _num_fmt_data(val)
                c.alignment = CENTER
                c.border    = THIN

                if ci >= cmp_col_start:
                    comp_idx              = ci - cmp_col_start
                    cmp_name              = compounds[comp_idx]
                    num_v, flag_cell, lod_cell = row_meta[comp_idx]
                    t_vals                = thresh_vals.get(cmp_name, {})
                    thresh_limit          = self._strictest(t_vals)
                    if thresh_limit is not None:
                        if (flag_cell not in ("ND", "<LOQ", "<")
                                and isinstance(num_v, (int, float))
                                and num_v > thresh_limit):
                            c.fill = ORANGE
                            c.font = Font(**FHE, bold=True)
                        # GREY + BOLD: threshold < LOD → false positive risk
                        elif flag_cell in ("ND", "<LOD", "<LOQ", "<"):
                            lod_num = (lod_cell if lod_cell is not None
                                       else (num_v if isinstance(num_v, (int, float)) else None))
                            if lod_num is not None and lod_num > thresh_limit:
                                c.fill = GRAY
                                c.font = _font(val, bold=True)
                                has_gray = True
            data_row += 1

        # ── Merge borehole column cells vertically ───────────────────────
        if has_depth and len(samples) > 1:
            # Walk sample rows and merge runs of the same borehole
            run_bh  = split_map[samples[0]][0]
            run_start = first_sample_row
            for idx, sid in enumerate(samples[1:], 1):
                bh = split_map[sid][0]
                row_num = first_sample_row + idx
                if bh != run_bh:
                    if row_num - 1 > run_start:   # >1 row → merge
                        ws.merge_cells(
                            start_row=run_start, start_column=1,
                            end_row=row_num - 1,   end_column=1
                        )
                        ws.cell(run_start, 1).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                    run_bh    = bh
                    run_start = row_num
            # Flush last run
            last_row = first_sample_row + len(samples) - 1
            if last_row > run_start:
                ws.merge_cells(
                    start_row=run_start, start_column=1,
                    end_row=last_row,     end_column=1
                )
                ws.cell(run_start, 1).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        self._write_legend(ws, data_row + 1, include_gray=has_gray)
        # ── Threshold source footnotes (only for keys with ≥1 defined value) ──
        active_keys = [k for k in thresh_keys
                       if any(thresh_vals.get(c, {}).get(k) is not None for c in compounds)]
        note_row = data_row + 2
        for note in self._threshold_source_notes(active_keys):
            ws.cell(row=note_row, column=1, value=f"* {note}").font = Font(
                **FEN, italic=True, color="808080")
            note_row += 1
        self._auto_width(ws, total_cols)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _thresh_keys(self, atype: str) -> list[str]:
        valid = set(ANALYSIS_THRESHOLDS.get(atype, []))
        if self.sel_thresh is not None:
            # Preserve user's ordering; drop keys not valid for this atype
            return [k for k in self.sel_thresh if k in valid]
        return ANALYSIS_THRESHOLDS.get(atype, [])

    @staticmethod
    def _strictest(t_vals: dict) -> float | None:
        vals = [v for v in t_vals.values() if v is not None]
        return min(vals) if vals else None

    def _write_header_row(self, ws, row_num: int, total_cols: int, hinfo: dict | None = None):
        if hinfo:
            parts = [
                ("שם פרויקט:", hinfo.get("project", "")),
                ("תאריך:",     hinfo.get("date", "")),
                ("מזמין:",     hinfo.get("client", "")),
            ]
            span = max(1, total_cols // len(parts))
            for i, (label, val) in enumerate(parts):
                col_start = i * span + 1
                col_end   = (i + 1) * span if i < len(parts) - 1 else total_cols
                ws.merge_cells(start_row=row_num, start_column=col_start,
                               end_row=row_num, end_column=col_end)
                c = ws.cell(row=row_num, column=col_start,
                            value=f"{label}  {val}")
                c.font      = Font(**FHE, bold=True)   # dark text, no fill
                c.alignment = WRAP_C
                c.border    = THIN
        else:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=total_cols)
            c = ws.cell(row=row_num, column=1,
                        value=f"{self.project}  |  {self.rep_date}  |  {self.client}")
            c.font      = Font(**FHE, bold=True)   # dark text, no fill
            c.alignment = WRAP_C
            c.border    = THIN

    @staticmethod
    def _threshold_source_notes(thresh_keys: list[str]) -> list[str]:
        """Return unique source footnote strings for the selected threshold keys, in order."""
        seen: set[str] = set()
        out: list[str] = []
        for k in thresh_keys:
            src = _THRESHOLD_SOURCES.get(k)
            if src and src not in seen:
                seen.add(src)
                out.append(src)
        return out

    @staticmethod
    def _write_legend(ws, start_row: int, include_gray: bool = True):
        items = [
            ("חריגה מערך סף",           ORANGE),
        ]
        if include_gray:
            items.append(("ערך הסף גדול מסף הגילוי", GRAY))
        for i, (label, fill) in enumerate(items):
            c = ws.cell(row=start_row + i, column=1, value=label)
            c.font   = Font(name="David", size=9, bold=True)
            c.fill   = fill
            c.border = THIN
            c.alignment = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def _auto_width(ws, n_cols: int, hdr_row: int = 2):
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 13
        for ci in range(3, n_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 13
        ws.row_dimensions[1].height = 20
        for r in range(2, hdr_row):
            ws.row_dimensions[r].height = 20   # metadata rows
        ws.row_dimensions[hdr_row].height = 28  # column header row


# ── Standalone helper: simple KTE groundwater BTEX report (A–H layout) ─────────

def _strip_ns(root: ET.Element) -> ET.Element:
    """Remove XML namespaces for easier searching (for SpreadsheetML .XLS)."""
    xml = ET.tostring(root, encoding="unicode")
    xml = re.sub(r'\s+xmlns(:\w+)?="[^"]*"', "", xml)
    xml = re.sub(r"<(\w+):", "<", xml)
    xml = re.sub(r"</(\w+):", "</", xml)
    xml = re.sub(r"(\s)(\w+):", r"\1", xml)
    return ET.fromstring(xml)


def build_kte_gw_btex_simple_from_xml(
    input_path: str | os.PathLike | bytes | io.BytesIO,
    output_path: str | os.PathLike | io.BytesIO,
) -> str | io.BytesIO:
    """
    Build a simple groundwater BTEX+MTBE sheet (A–H) from
    KTE 'Client GROUNDWATER - 1' SpreadsheetML XML report.

    input_path  — file path, raw bytes, or BytesIO
    output_path — file path or BytesIO
    """
    if isinstance(input_path, (bytes, bytearray)):
        raw = input_path
    elif isinstance(input_path, io.BytesIO):
        raw = input_path.read()
    else:
        raw = Path(input_path).read_bytes()
    root = ET.fromstring(raw)
    root = _strip_ns(root)

    # Find "Client GROUNDWATER - 1" worksheet
    ws_xml = None
    for w in root.findall(".//Worksheet"):
        name = (w.get("Name") or "").strip().lower()
        if "client groundwater - 1" in name:
            ws_xml = w
            break
    if ws_xml is None:
        raise RuntimeError("לא נמצא worksheet בשם 'Client GROUNDWATER - 1'")

    table = ws_xml.find(".//Table")
    if table is None:
        raise RuntimeError("לא נמצא <Table> ב‑Client GROUNDWATER - 1")

    # Convert XML rows → list[list[str]]
    rows: list[list[str]] = []
    for row_el in table.findall("Row"):
        cells: list[str] = []
        prev_idx = 0
        for cell_el in row_el.findall("Cell"):
            idx_attr = cell_el.get("Index")
            if idx_attr is not None:
                idx = int(idx_attr)
                gap = idx - 1 - prev_idx
                if gap > 0:
                    cells.extend([""] * gap)
            data_el = cell_el.find("Data")
            val = data_el.text if (data_el is not None and data_el.text) else ""
            cells.append(val.strip())
            prev_idx = len(cells) - 1
        rows.append(cells)

    # Locate sample IDs and sampling dates
    sample_row = date_row = None
    for r in rows:
        joined = " ".join(r).lower()
        if "client sample id" in joined:
            sample_row = r
        if "client sampling date" in joined:
            date_row = r
    if sample_row is None or date_row is None:
        raise RuntimeError("לא מצאתי שורות Client Sample ID / Client Sampling Date")

    # In this SpreadsheetML export the data starts at column index 4 (E in Excel)
    well_names = sample_row[4:]
    dates = date_row[4:]

    def _find_param_row(name_substr: str, preferred_unit: str | None = None) -> list[str] | None:
        """
        Find first row whose first cell contains name_substr (case-insensitive).
        If preferred_unit is given, prefer a row whose Unit column matches it.
        """
        key = name_substr.lower()
        candidates: list[list[str]] = []
        for r in rows:
            if r and key in r[0].lower():
                candidates.append(r)
        if not candidates:
            return None
        if preferred_unit is None:
            return candidates[0]
        for r in candidates:
            if len(r) > 2 and (r[2] or "").strip() == preferred_unit:
                return r
        return candidates[0]

    r_benzene = _find_param_row("benzene")
    r_toluene = _find_param_row("toluene")
    r_ethylbenz = _find_param_row("ethylbenzene")
    r_meta_para = _find_param_row("meta- & para-xylene")
    r_ortho_xyl = _find_param_row("ortho-xylene")
    r_mtbe = _find_param_row("methyl tert-butyl ether (mtbe)")

    if not all([r_benzene, r_toluene, r_ethylbenz, r_meta_para, r_ortho_xyl, r_mtbe]):
        raise RuntimeError("לא נמצאו כל שורות BTEX/MTBE ב‑INPUT")

    def _vals_ugL_to_mgL(r: list[str]) -> list[float | str]:
        unit = (r[2] or "").strip()  # col C = Unit
        factor = 0.001 if unit == "µg/L" else 1.0
        out: list[float | str] = []
        for v in r[4:]:
            v = v.strip()
            if not v:
                out.append("")
            elif v.startswith("<"):
                try:
                    num = float(v[1:])
                    out.append(f"<{round(num * factor, 4)}")
                except ValueError:
                    out.append(v)
            else:
                try:
                    num = float(v)
                    out.append(round(num * factor, 4))
                except ValueError:
                    out.append(v)
        return out

    benzene_vals = _vals_ugL_to_mgL(r_benzene)
    toluene_vals = _vals_ugL_to_mgL(r_toluene)
    ethyl_vals = _vals_ugL_to_mgL(r_ethylbenz)
    meta_para_vals = _vals_ugL_to_mgL(r_meta_para)
    ortho_vals = _vals_ugL_to_mgL(r_ortho_xyl)
    mtbe_vals = _vals_ugL_to_mgL(r_mtbe)

    # Xylene = sum(meta+para, ortho)
    xylene_vals: list[float | str] = []
    for mp, o in zip(meta_para_vals, ortho_vals):
        try:
            mp_num = float(str(mp).lstrip("<"))
            o_num = float(str(o).lstrip("<"))
            xylene_vals.append(round(mp_num + o_num, 4))
        except ValueError:
            xylene_vals.append("")

    # Build simple Excel sheet
    wb = openpyxl.Workbook()

    # ── Sheet 1: BTEX + MTBE ─────────────────────────────────────────────
    ws_out = wb.active
    ws_out.title = "מי תהום BTEX"
    ws_out.sheet_view.rightToLeft = True

    NUM_FMT = "0.0000"   # 4 decimal places for all numeric data cells

    def _set_cell(r: int, c: int, val, bold: bool = False, num: bool = False):
        cell = ws_out.cell(row=r, column=c, value=val)
        cell.font = Font(**FHE, bold=bold)   # David 9 everywhere
        cell.alignment = CENTER
        cell.border = THIN
        if num and val not in (None, ""):
            cell.number_format = NUM_FMT
        return cell

    # Row 1: headers (all David 9 bold)
    _set_cell(1, 1, "שם קידוח", bold=True)
    _set_cell(1, 2, "תאריך דיגום", bold=True)
    _set_cell(1, 3, "", bold=True)
    _set_cell(1, 4, "בנזן", bold=True)
    _set_cell(1, 5, "טולואן", bold=True)
    _set_cell(1, 6, "אתיל בנזן", bold=True)
    _set_cell(1, 7, "כסילן", bold=True)
    # MTBE is English → Times New Roman 9 bold
    _c = ws_out.cell(row=1, column=8, value="MTBE")
    _c.font = Font(**FEN, bold=True)
    _c.alignment = CENTER
    _c.border = THIN

    # Row 2: units  (A,B empty | C יחידות | D-H מ"ג/ליטר)
    _set_cell(2, 1, "")
    _set_cell(2, 2, "")
    _set_cell(2, 3, "יחידות", bold=True)
    for col in range(4, 9):
        _set_cell(2, col, 'מ"ג/ליטר')

    # Row 3: restoration targets  (A,B empty | C label | D-H values David 9 + 4dp)
    _set_cell(3, 1, "")
    _set_cell(3, 2, "")
    _set_cell(3, 3, "ערכי יעד לשיקום ^", bold=True)
    targets = {
        4: 0.094,   # Benzene
        5: 13.0,    # Toluene
        6: 5.6,     # Ethylbenzene
        7: 9.4,     # Xylenes
        8: 0.75,    # MTBE
    }
    for col in range(4, 9):
        _set_cell(3, col, targets.get(col, ""), num=True)

    def _norm_well(name: str) -> str:
        """MT-x → מת-x"""
        s = (name or "").strip()
        if s.upper().startswith("MT-"):
            return "מת-" + s[3:]
        return s

    # Sample rows (rows 4+)
    for idx, (well, dt) in enumerate(zip(well_names, dates), start=4):
        # well name: Hebrew → David 9, English → Times 8 (auto-detect via _font)
        _wn = _norm_well(well)
        _c = ws_out.cell(row=idx, column=1, value=_wn)
        _c.font = _font(_wn)
        _c.alignment = CENTER
        _c.border = THIN
        _set_cell(idx, 2, dt or "")           # date      — David 9
        _set_cell(idx, 3, "")

        def _put(col: int, arr: list[float | str]):
            pos = idx - 4
            if 0 <= pos < len(arr):
                val = arr[pos]
                cell = _set_cell(idx, col, val, num=True)   # David 9 + 4dp
                t = targets.get(col)
                if t is not None:
                    try:
                        num_val = float(str(val).lstrip("<"))
                        if num_val > t:
                            cell.fill = GRAY
                            cell.font = Font(**FHE, bold=True)
                    except (ValueError, TypeError):
                        pass

        _put(4, benzene_vals)
        _put(5, toluene_vals)
        _put(6, ethyl_vals)
        _put(7, xylene_vals)
        _put(8, mtbe_vals)

    # Basic widths
    ws_out.column_dimensions["A"].width = 18
    ws_out.column_dimensions["B"].width = 14
    ws_out.column_dimensions["C"].width = 18
    for col in range(4, 9):
        ws_out.column_dimensions[get_column_letter(col)].width = 12

    # ── Sheet 2: Chlorite/Chlorate/Chloride/Perchlorate ─────────────────
    # Parameters:
    #   D: כלוריט + כלוראט  → "Sum of chlorites and chlorates"
    #   E: כלוריד           → "Chloride" (mg/L row)
    #   F: פרכלורט          → "Perchlorate"
    #   G: כלוראט           → "Chlorate"

    r_sum_chlor = _find_param_row("sum of chlorites and chlorates")
    r_chloride  = _find_param_row("chloride", preferred_unit="mg/L")
    r_percl     = _find_param_row("perchlorate")
    r_chlorate  = _find_param_row("chlorate")

    def _save_wb(wb, dest):
        if isinstance(dest, io.BytesIO):
            wb.save(dest)
            dest.seek(0)
            return dest
        os.makedirs(os.path.dirname(str(dest)) or ".", exist_ok=True)
        wb.save(dest)
        return str(dest)

    if any(r is None for r in (r_sum_chlor, r_chloride, r_percl, r_chlorate)):
        return _save_wb(wb, output_path)

    def _vals_to_mgL(r: list[str]) -> list[float | str]:
        unit = (r[2] or "").strip()
        factor = 0.001 if unit == "µg/L" else 1.0
        out: list[float | str] = []
        for v in r[4:]:
            v = v.strip()
            if not v:
                out.append("")
            elif v.startswith("<"):
                try:
                    num = float(v[1:])
                    out.append(f"<{round(num * factor, 4)}")
                except ValueError:
                    out.append(v)
            else:
                try:
                    num = float(v)
                    out.append(round(num * factor, 4))
                except ValueError:
                    out.append(v)
        return out

    sum_chlor_vals = _vals_to_mgL(r_sum_chlor)
    chloride_vals  = _vals_to_mgL(r_chloride)
    percl_vals     = _vals_to_mgL(r_percl)
    chlorate_vals  = _vals_to_mgL(r_chlorate)

    ws2 = wb.create_sheet(title="מי שתייה – כלורידים")
    ws2.sheet_view.rightToLeft = True

    def _set2(r: int, c: int, val, bold: bool = False, he: bool = True):
        cell = ws2.cell(row=r, column=c, value=val)
        if he:
            cell.font = Font(**FHE, bold=bold)
        else:
            base = FEN.copy()
            base["bold"] = bold
            cell.font = Font(**base)
        cell.alignment = CENTER
        cell.border = THIN
        return cell

    # Row 1: headers
    _set2(1, 1, "שם קידוח", bold=True)
    _set2(1, 2, "תאריך דיגום", bold=True)
    _set2(1, 3, "", bold=True)
    _set2(1, 4, "כלוריט + כלוראט", bold=True)
    _set2(1, 5, "כלוריד", bold=True)
    _set2(1, 6, "פרכלורט", bold=True)
    _set2(1, 7, "כלוראט", bold=True)

    # Row 2: units
    _set2(2, 1, "")
    _set2(2, 2, "")
    _set2(2, 3, "יחידות", bold=True)
    for col in range(4, 8):
        _set2(2, col, 'מ"ג/ליטר')

    # Row 3: drinking water thresholds
    _set2(3, 1, "")
    _set2(3, 2, "")
    _set2(3, 3, "ערך סף למי שתייה^", bold=True)
    _set2(3, 4, "--")
    _set2(3, 5, 400)   # כלוריד
    _set2(3, 6, "--")
    _set2(3, 7, "--")

    # Sample rows
    for idx, (well, dt) in enumerate(zip(well_names, dates), start=4):
        _set2(idx, 1, well or "")
        _set2(idx, 2, dt or "")
        _set2(idx, 3, "")

        def _put2(col: int, arr: list[float | str]):
            pos = idx - 4
            if 0 <= pos < len(arr):
                _set2(idx, col, arr[pos])

        _put2(4, sum_chlor_vals)
        _put2(5, chloride_vals)
        _put2(6, percl_vals)
        _put2(7, chlorate_vals)

    # Column widths
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 22
    for col in range(4, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    return _save_wb(wb, output_path)
