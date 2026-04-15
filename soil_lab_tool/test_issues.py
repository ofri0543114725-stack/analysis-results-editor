"""Investigate all reported issues."""
import sys, io, os
sys.path.insert(0, os.path.dirname(__file__))
from parsers.soil.machon_haneft import MachonHaneftSoilParser
from core.excel_output import LabReportExcel
from core.threshold_manager import ThresholdManager
import openpyxl

BASE = r"C:\Users\Asaf\claude\laboratory_results_analsys\Laboratory_results\machon_haneft\soil"
files = [os.path.join(BASE, "shimshon_1.xlsx"), os.path.join(BASE, "shishon_2.xlsx")]

all_records = []
for path in files:
    with open(path, "rb") as f:
        data = f.read()
    parser = MachonHaneftSoilParser()
    recs = parser.parse(io.BytesIO(data))
    all_records.extend(recs)

# Check TPH LOD values
tph = [r for r in all_records if r["analysis_type"] == "SOIL_TPH"]
print("TPH LOD values sample:")
seen = set()
for r in tph[:10]:
    k = r["compound"]
    if k not in seen:
        seen.add(k)
        print(f"  compound={r['compound']} lod={r.get('lod')} flag={r['flag']} value={r['value']}")

# Check TIER1_IND_SOIL_HM_6 for C10-C40
THRESH_DIR  = os.path.join(os.path.dirname(__file__), "thresholds")
MAIN_THRESH = os.path.join(THRESH_DIR, "soil_vsl_tier1_v7_2024.xlsx")
VSL_FULL    = os.path.join(THRESH_DIR, "soil_vsl_v7_full.xlsx")
tm = ThresholdManager(MAIN_THRESH, vsl_full_path=VSL_FULL if os.path.exists(VSL_FULL) else None)

print("\nTIER thresholds for C10-C40:")
for key in ["VSL_SOIL", "TIER1_RES_SOIL_VH", "TIER1_RES_SOIL_HM_0_6", "TIER1_RES_SOIL_HM_6",
            "TIER1_IND_SOIL_VH", "TIER1_IND_SOIL_HM_0_6", "TIER1_IND_SOIL_HM_6", "TIER1_IND_SOIL_LOW"]:
    v = tm.get_threshold("C10-C40", key)
    print(f"  {key}: {v}")

# Check with actual CAS for diesel range:
print("\nFor DRO CAS=C10-C40 exactly:")
v = tm.get_threshold("C10-C40", "TIER1_IND_SOIL_HM_6")
print(f"  TIER1_IND_SOIL_HM_6 = {v}")

# Build Excel with TIER1_IND_SOIL_HM_6 selected
out_buf = io.BytesIO()
builder = LabReportExcel(
    records=all_records, threshold_manager=tm, output_path=out_buf,
    project_name="Test",
    selected_thresholds=["VSL_SOIL", "TIER1_IND_SOIL_HM_6"],
    combine_tph_mbtex=False,
)
builder.build()
out_buf.seek(0)

wb = openpyxl.load_workbook(io.BytesIO(out_buf.read()))
ws = wb["קרקע TPH"]
print("\nTPH sheet first 10 rows:")
for row_idx in range(1, 11):
    row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 8)]
    # Also check number_format of row 5
    if row_idx == 5:
        fmts = [ws.cell(row=row_idx, column=c).number_format for c in range(4, 7)]
        print(f"  row {row_idx} D-F formats: {fmts}")
    print(f"  row {row_idx}: {row_vals}")

print("\nTPH data rows (rows 7-12):")
for row_idx in range(7, 13):
    row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 8)]
    lod_info = [ws.cell(row=row_idx, column=c).value for c in range(4, 7)]
    print(f"  row {row_idx}: {row_vals}")
